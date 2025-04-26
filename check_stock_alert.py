import mysql.connector
from apscheduler.schedulers.background import BackgroundScheduler
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
import logging
from datetime import datetime, timedelta
import os
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.drawing.image import Image
from openpyxl.styles import NamedStyle

# ======================
# CONFIGURATION
# ======================

DB_CONFIG = {
    'host': 'localhost',
    'user': 'root',
    'password': '',
    'database': 'it_park',
    'charset': 'utf8mb4'
}

EMAIL_CONFIG = {
    'smtp_server': 'smtp.gmail.com',
    'smtp_port': 587,
    'email': 'douha.hm17@gmail.com',
    'password': 'iwsewbpeawxkcfod'
}

COMPANY_LOGO = "company_logo.png"

# ======================
# CORE FUNCTIONALITY
# ======================

def setup_logging():
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler('purchase_requests.log'),
            logging.StreamHandler()
        ]
    )
    return logging.getLogger(__name__)

logger = setup_logging()

def get_all_user_emails(cursor):
    """Retrieve all approved user emails from the database"""
    try:
        cursor.execute("""
            SELECT email FROM users 
            WHERE status = 'Approuvé' 
            AND email IS NOT NULL
            AND email != ''
        """)
        emails = [row['email'] for row in cursor.fetchall()]
        logger.info(f"Found {len(emails)} approved users")
        return emails
    except mysql.connector.Error as err:
        logger.error(f"Erreur lors de la récupération des emails: {err}")
        return []

def generate_purchase_request(items, cursor):
    """Generate clean purchase request with white background and no borders"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Demande d'Achat"
    
    # Add company logo if available
    if os.path.exists(COMPANY_LOGO):
        try:
            img = Image(COMPANY_LOGO)
            img.width = 200
            img.height = 100
            ws.add_image(img, 'A1')
        except Exception as e:
            logger.warning(f"Could not add company logo: {str(e)}")
    
    # Create clean styles
    header_style = NamedStyle(name="header_style")
    header_style.font = Font(bold=True, size=12)
    header_style.alignment = Alignment(horizontal='left')
    
    title_style = NamedStyle(name="title_style")
    title_style.font = Font(bold=True, size=14)
    title_style.alignment = Alignment(horizontal='center')
    
    table_header_style = NamedStyle(name="table_header_style")
    table_header_style.font = Font(bold=True)
    table_header_style.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add header information (all editable)
    ws.merge_cells('A5:D5')
    ws['A5'] = "Objet : Demande d'achat de matériel informatique"
    ws['A5'].style = title_style
    
    ws['A7'] = "À : Service des Achats"
    ws['A7'].style = header_style
    
    ws['A8'] = "De : Service Gestion de Stock"
    ws['A8'].style = header_style
    
    ws['A9'] = f"Date : {datetime.now().strftime('%d/%m/%Y')}"
    ws['A9'].style = header_style
    
    ws['A10'] = f"Réf. : DA-{datetime.now().strftime('%Y%m%d')}-001"
    ws['A10'].style = header_style
    
    # Add introduction text (editable)
    ws.merge_cells('A12:D12')
    ws['A12'] = "Madame, Monsieur,"
    ws['A12'].style = header_style
    
    ws.merge_cells('A13:D13')
    ws['A13'] = "Par la présente, je souhaite soumettre une demande d'achat pour le matériel suivant,"
    ws['A13'].style = header_style
    
    ws.merge_cells('A14:D14')
    ws['A14'] = "nécessaire au bon fonctionnement de notre service :"
    ws['A14'].style = header_style
    
    # Add table headers
    headers = [
        'Désignation',
        'Sous-catégorie',
        'Quantité à Commander', 
        'Commentaires'
    ]
    
    ws.append([])  # Empty row before table
    ws.append(headers)
    
    # Apply clean table header styles
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=17, column=col)
        cell.style = table_header_style
    
    # Add only items that need purchasing (quantity < threshold)
    items_to_order = [item for item in items if item['quantite'] < item['seuil_min']]
    
    if not items_to_order:
        logger.info("Aucun équipement ne nécessite de commande")
        return None
    
    for idx, item in enumerate(items_to_order, start=1):
        qty_to_order = item['seuil_min'] - item['quantite']
        
        row = [
            item.get('article', item['sous_categorie']),  # Désignation
            item['sous_categorie'],                      # Sous-catégorie
            qty_to_order,                                # Quantité à Commander
            ""                                           # Commentaires
        ]
        ws.append(row)
    
    # Add data validation for quantity (must be positive number)
    dv_qty = DataValidation(type="whole", operator="greaterThanOrEqual", formula1="1")
    dv_qty.add(f"C18:C{17+len(items_to_order)}")
    ws.add_data_validation(dv_qty)
    
    # Add motive section
    motive_row = 17 + len(items_to_order) + 2
    ws.merge_cells(f'A{motive_row}:D{motive_row}')
    ws[f'A{motive_row}'] = "Motif : Réapprovisionnement du stock selon les seuils minimums définis"
    ws[f'A{motive_row}'].style = header_style
    
    # Add closing text
    closing_row = motive_row + 2
    ws.merge_cells(f'A{closing_row}:D{closing_row}')
    ws[f'A{closing_row}'] = "Je vous prie de bien vouloir valider cette demande et lancer le processus d'achat dans les meilleurs délais."
    ws[f'A{closing_row}'].style = header_style
    
    # Add signature section
    signature_row = closing_row + 3
    ws.merge_cells(f'A{signature_row}:D{signature_row}')
    ws[f'A{signature_row}'] = "Cordialement,"
    ws[f'A{signature_row}'].style = header_style
    
    ws.merge_cells(f'A{signature_row+1}:D{signature_row+1}')
    ws[f'A{signature_row+1}'] = "Service Gestion de Stock"
    ws[f'A{signature_row+1}'].style = header_style
    
    # Format columns
    column_widths = {
        'A': 30,  # Désignation
        'B': 20,  # Sous-catégorie
        'C': 20,  # Quantité à Commander
        'D': 30   # Commentaires
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Format quantity column
    for cell in ws['C'][17:]:  # Column C (Quantité à Commander)
        if cell.row >= 18:
            cell.number_format = '#,##0'
    
    filename = f"Demande_Achat_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    return filename

def send_purchase_request(recipients, excel_file):
    """Send the purchase request with clean instructions"""
    try:
        msg = MIMEMultipart()
        msg['From'] = EMAIL_CONFIG['email']
        msg['Subject'] = f"📋 Demande d'Achat - {datetime.now().strftime('%d/%m/%Y')}"
        
        body = f"""
        Bonjour,
        
        Veuillez trouver ci-joint la demande d'achat pour les équipements nécessitant un réapprovisionnement.
        
        Tous les champs sont modifiables. Merci de :
        1. Vérifier les quantités proposées
        2. Ajouter des commentaires si nécessaire
        3. Retourner ce fichier complété
        
        Date limite de traitement: {(datetime.now() + timedelta(days=3)).strftime('%d/%m/%Y')}
        
        Cordialement,
        Service Gestion de Stock
        """
        msg.attach(MIMEText(body, 'plain', 'utf-8'))
        
        with open(excel_file, "rb") as f:
            attach = MIMEApplication(f.read(), _subtype="xlsx")
            attach.add_header(
                'Content-Disposition', 
                'attachment', 
                filename=os.path.basename(excel_file)
            )
            msg.attach(attach)
        
        with smtplib.SMTP(EMAIL_CONFIG['smtp_server'], EMAIL_CONFIG['smtp_port']) as server:
            server.starttls()
            server.login(EMAIL_CONFIG['email'], EMAIL_CONFIG['password'])
            
            for email in recipients:
                msg['To'] = email
                server.sendmail(EMAIL_CONFIG['email'], email, msg.as_string())
                logger.info(f"Demande d'achat envoyée à {email}")
                
    except Exception as e:
        logger.error(f"Erreur d'envoi: {str(e)}")
    finally:
        if os.path.exists(excel_file):
            os.remove(excel_file)


def check_and_generate_orders():
    """Main function to check stock and generate orders"""
    try:
        conn = mysql.connector.connect(**DB_CONFIG)
        cursor = conn.cursor(dictionary=True)
        
        # Get only items that need purchasing (quantity < threshold)
        cursor.execute("""
            SELECT e.article, s.sous_categorie, s.quantite, s.seuil_min 
            FROM stock s
            LEFT JOIN equipements e ON s.sous_categorie = e.sous_categorie
            WHERE s.quantite < s.seuil_min
            GROUP BY s.sous_categorie
            ORDER BY s.quantite/s.seuil_min ASC
        """)
        items = cursor.fetchall()
        
        if items:
            logger.info(f"Génération de demande d'achat pour {len(items)} articles nécessitant réapprovisionnement")
            excel_file = generate_purchase_request(items, cursor)
            
            if excel_file:  # Only send if file was generated
                recipients = get_all_user_emails(cursor)
                if recipients:
                    send_purchase_request(recipients, excel_file)
                else:
                    logger.error("Aucun utilisateur approuvé trouvé pour recevoir les alertes")
        else:
            logger.info("Aucun équipement ne nécessite de commande")
            
    except mysql.connector.Error as err:
        logger.error(f"Erreur base de données: {err}")
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close()
            conn.close()

# ======================
# SCHEDULER & MAIN
# ======================

if __name__ == '__main__':
    # First run immediately for testing
    check_and_generate_orders()
    
    # Then schedule regular checks
    scheduler = BackgroundScheduler()
    scheduler.add_job(
        check_and_generate_orders,
        'interval',
        minutes=1,  # Check every minute for testing
        timezone='UTC'
    )
    scheduler.start()
    
    try:
        while True:
            pass
    except (KeyboardInterrupt, SystemExit):
        scheduler.shutdown()